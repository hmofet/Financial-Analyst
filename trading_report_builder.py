#!/usr/bin/env python3
"""
Trading Report Builder - Portable Version
A Windows Desktop Application without matplotlib DLL issues.
Uses tkinter Canvas for charts instead.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import sys
import webbrowser
import tempfile

# For exports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class SimpleBarChart(tk.Canvas):
    """Simple bar chart using tkinter Canvas - no matplotlib needed"""
    
    def __init__(self, parent, width=600, height=400, **kwargs):
        super().__init__(parent, width=width, height=height, bg='white', **kwargs)
        self.width = width
        self.height = height
        self.margin = {'left': 80, 'right': 20, 'top': 40, 'bottom': 60}
        
    def clear(self):
        self.delete('all')
        
    def draw_bar_chart(self, data, labels, title="", colors=None, horizontal=False):
        """Draw a bar chart
        data: list of values
        labels: list of labels
        title: chart title
        colors: list of colors or single color
        horizontal: if True, draw horizontal bars
        """
        self.clear()
        
        if not data or not labels:
            self.create_text(self.width/2, self.height/2, text="No data available", 
                           font=('Segoe UI', 12))
            return
            
        # Chart area
        chart_left = self.margin['left']
        chart_right = self.width - self.margin['right']
        chart_top = self.margin['top']
        chart_bottom = self.height - self.margin['bottom']
        chart_width = chart_right - chart_left
        chart_height = chart_bottom - chart_top
        
        # Title
        self.create_text(self.width/2, 15, text=title, font=('Segoe UI', 12, 'bold'))
        
        # Normalize data
        max_val = max(abs(v) for v in data) if data else 1
        min_val = min(0, min(data)) if data else 0
        
        if horizontal:
            # Horizontal bars
            bar_height = chart_height / len(data) * 0.7
            gap = chart_height / len(data) * 0.3
            
            for i, (val, label) in enumerate(zip(data, labels)):
                y = chart_top + i * (bar_height + gap) + gap/2
                
                # Color
                if colors:
                    color = colors[i] if isinstance(colors, list) else colors
                else:
                    color = '#4CAF50' if val >= 0 else '#F44336'
                
                # Bar
                bar_width = abs(val) / max_val * chart_width * 0.8
                if val >= 0:
                    self.create_rectangle(chart_left, y, chart_left + bar_width, 
                                        y + bar_height, fill=color, outline='')
                else:
                    self.create_rectangle(chart_left, y, chart_left + bar_width,
                                        y + bar_height, fill=color, outline='')
                
                # Label
                self.create_text(chart_left - 5, y + bar_height/2, text=str(label)[:12],
                               anchor='e', font=('Segoe UI', 8))
                
                # Value
                val_text = f"${val:,.0f}" if abs(val) > 1 else f"{val:.2f}"
                self.create_text(chart_left + bar_width + 5, y + bar_height/2,
                               text=val_text, anchor='w', font=('Segoe UI', 8))
        else:
            # Vertical bars
            bar_width = chart_width / len(data) * 0.7
            gap = chart_width / len(data) * 0.3
            
            # Zero line
            if min_val < 0:
                zero_y = chart_bottom - (0 - min_val) / (max_val - min_val) * chart_height
                self.create_line(chart_left, zero_y, chart_right, zero_y, fill='black')
            else:
                zero_y = chart_bottom
            
            for i, (val, label) in enumerate(zip(data, labels)):
                x = chart_left + i * (bar_width + gap) + gap/2
                
                # Color
                if colors:
                    color = colors[i] if isinstance(colors, list) else colors
                else:
                    color = '#4CAF50' if val >= 0 else '#F44336'
                
                # Bar height
                bar_height = abs(val) / max_val * chart_height * 0.8
                
                if val >= 0:
                    self.create_rectangle(x, zero_y - bar_height, x + bar_width,
                                        zero_y, fill=color, outline='')
                else:
                    self.create_rectangle(x, zero_y, x + bar_width,
                                        zero_y + bar_height, fill=color, outline='')
                
                # Label (rotated approximation)
                label_text = str(label)[:8]
                self.create_text(x + bar_width/2, chart_bottom + 10, text=label_text,
                               anchor='n', font=('Segoe UI', 7), angle=45)
                
    def draw_pie_chart(self, data, labels, title="", colors=None):
        """Draw a simple pie chart"""
        self.clear()
        
        if not data or not labels or sum(data) == 0:
            self.create_text(self.width/2, self.height/2, text="No data available",
                           font=('Segoe UI', 12))
            return
            
        # Title
        self.create_text(self.width/2, 15, text=title, font=('Segoe UI', 12, 'bold'))
        
        # Pie chart center and radius
        cx = self.width / 2 - 80
        cy = self.height / 2 + 10
        radius = min(self.width, self.height) / 3
        
        # Default colors
        default_colors = ['#4CAF50', '#2196F3', '#FF9800', '#9C27B0', '#F44336',
                         '#00BCD4', '#FFEB3B', '#795548', '#607D8B', '#E91E63']
        
        total = sum(data)
        start_angle = 0
        
        for i, (val, label) in enumerate(zip(data, labels)):
            # Calculate angle
            angle = val / total * 360
            
            # Color
            if colors:
                color = colors[i] if isinstance(colors, list) else colors
            else:
                color = default_colors[i % len(default_colors)]
            
            # Draw arc
            self.create_arc(cx - radius, cy - radius, cx + radius, cy + radius,
                          start=start_angle, extent=angle, fill=color, outline='white')
            
            # Legend
            legend_y = 60 + i * 20
            self.create_rectangle(self.width - 150, legend_y, self.width - 135, 
                                legend_y + 12, fill=color)
            pct = val / total * 100
            self.create_text(self.width - 130, legend_y + 6, 
                           text=f"{label[:10]} ({pct:.1f}%)", anchor='w',
                           font=('Segoe UI', 8))
            
            start_angle += angle


class TradingReportBuilder:
    def __init__(self, root):
        self.root = root
        self.root.title("Trading Report Builder - Questrade Analysis")
        self.root.geometry("1400x900")
        self.root.minsize(1200, 700)
        
        # Data storage
        self.transactions_df = None
        self.trades_df = None
        self.dividends_df = None
        self.fifo_results = None
        self.stock_summary = None
        
        # Category definitions
        self.categories = {
            "TSX Mining": ["ABX.TO", "CCO.TO", "TECK-B.TO", "NTR.TO", "FM.TO", "FNV.TO", 
                          "AGI.TO", "AEM.TO", "K.TO", "WPM.TO", "LUN.TO", "IVN.TO", 
                          "NXE.TO", "CS.TO", "B2GOLD.TO"],
            "Dividend": ["ENB.TO", "SU.TO", "BCE.TO", "JNJ", "ABBV", "PFE", "KO", "PG", 
                        "T.TO", "BNS.TO"],
            "Tech": ["AAPL", "MSFT", "NVDA", "GOOGL", "META", "AMZN", "TSLA", "AMD", 
                    "CRM", "SHOP.TO", "ADBE", "INTC", "CSCO", "ORCL"],
            "Blue Chip": ["JPM", "WMT", "V", "UNH", "LLY", "MRK", "BMY", "CAT", "HD", 
                         "MA", "DIS", "XOM", "CVX"]
        }
        
        self.setup_styles()
        self.create_menu()
        self.create_main_layout()
        self.create_status_bar()
        
    def setup_styles(self):
        """Configure ttk styles"""
        style = ttk.Style()
        style.theme_use('clam')
        
        style.configure('Header.TLabel', font=('Segoe UI', 12, 'bold'))
        style.configure('Title.TLabel', font=('Segoe UI', 14, 'bold'))
        style.configure('Action.TButton', font=('Segoe UI', 10))
        style.configure('Treeview', font=('Consolas', 9), rowheight=25)
        style.configure('Treeview.Heading', font=('Segoe UI', 10, 'bold'))
        
    def create_menu(self):
        """Create application menu"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Import Excel...", command=self.import_excel, accelerator="Ctrl+O")
        file_menu.add_command(label="Import CSV...", command=self.import_csv)
        file_menu.add_separator()
        file_menu.add_command(label="Export to Excel...", command=self.export_excel, accelerator="Ctrl+E")
        file_menu.add_command(label="Export to PDF...", command=self.export_pdf)
        file_menu.add_command(label="Export to HTML...", command=self.export_html)
        file_menu.add_separator()
        file_menu.add_command(label="Print Report...", command=self.print_report, accelerator="Ctrl+P")
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        
        # View menu
        view_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        view_menu.add_command(label="Raw Transactions", command=lambda: self.show_tab(0))
        view_menu.add_command(label="Trades Analysis", command=lambda: self.show_tab(1))
        view_menu.add_command(label="Dividends", command=lambda: self.show_tab(2))
        view_menu.add_command(label="P&L Summary", command=lambda: self.show_tab(3))
        view_menu.add_command(label="Charts", command=lambda: self.show_tab(4))
        
        # Reports menu
        reports_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Reports", menu=reports_menu)
        reports_menu.add_command(label="Top 10 Gainers", command=self.show_top_gainers)
        reports_menu.add_command(label="Top 10 Losers", command=self.show_top_losers)
        reports_menu.add_command(label="By Category", command=self.show_by_category)
        reports_menu.add_command(label="Monthly Summary", command=self.show_monthly_summary)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Help", menu=help_menu)
        help_menu.add_command(label="About", command=self.show_about)
        
        # Keyboard shortcuts
        self.root.bind('<Control-o>', lambda e: self.import_excel())
        self.root.bind('<Control-e>', lambda e: self.export_excel())
        self.root.bind('<Control-p>', lambda e: self.print_report())
        
    def create_main_layout(self):
        """Create main application layout"""
        # Main container
        main_container = ttk.Frame(self.root)
        main_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Left panel - Filters
        self.create_filter_panel(main_container)
        
        # Right panel - Content
        self.content_frame = ttk.Frame(main_container)
        self.content_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Notebook for tabs
        self.notebook = ttk.Notebook(self.content_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Create tabs
        self.create_raw_data_tab()
        self.create_trades_tab()
        self.create_dividends_tab()
        self.create_pnl_tab()
        self.create_charts_tab()
        
    def create_filter_panel(self, parent):
        """Create filter panel on the left"""
        filter_frame = ttk.LabelFrame(parent, text="Filters", padding=10)
        filter_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        
        # Date range
        ttk.Label(filter_frame, text="Date Range:", style='Header.TLabel').pack(anchor=tk.W, pady=(0, 5))
        
        date_frame = ttk.Frame(filter_frame)
        date_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(date_frame, text="From:").pack(anchor=tk.W)
        self.date_from = ttk.Entry(date_frame, width=15)
        self.date_from.pack(fill=tk.X)
        self.date_from.insert(0, "2025-01-01")
        
        ttk.Label(date_frame, text="To:").pack(anchor=tk.W, pady=(5, 0))
        self.date_to = ttk.Entry(date_frame, width=15)
        self.date_to.pack(fill=tk.X)
        self.date_to.insert(0, "2025-12-31")
        
        # Category filter
        ttk.Label(filter_frame, text="Category:", style='Header.TLabel').pack(anchor=tk.W, pady=(10, 5))
        
        self.category_var = tk.StringVar(value="All")
        categories = ["All", "TSX Mining", "Dividend", "Tech", "Blue Chip", "Other"]
        self.category_combo = ttk.Combobox(filter_frame, textvariable=self.category_var, 
                                           values=categories, state="readonly", width=18)
        self.category_combo.pack(fill=tk.X)
        
        # Action filter
        ttk.Label(filter_frame, text="Action:", style='Header.TLabel').pack(anchor=tk.W, pady=(10, 5))
        
        self.action_var = tk.StringVar(value="All")
        self.action_combo = ttk.Combobox(filter_frame, textvariable=self.action_var,
                                         values=["All", "Buy", "Sell", "DIV"], state="readonly", width=18)
        self.action_combo.pack(fill=tk.X)
        
        # Currency filter
        ttk.Label(filter_frame, text="Currency:", style='Header.TLabel').pack(anchor=tk.W, pady=(10, 5))
        
        self.currency_var = tk.StringVar(value="All")
        self.currency_combo = ttk.Combobox(filter_frame, textvariable=self.currency_var,
                                           values=["All", "CAD", "USD"], state="readonly", width=18)
        self.currency_combo.pack(fill=tk.X)
        
        # Symbol search
        ttk.Label(filter_frame, text="Symbol:", style='Header.TLabel').pack(anchor=tk.W, pady=(10, 5))
        self.symbol_var = tk.StringVar()
        self.symbol_entry = ttk.Entry(filter_frame, textvariable=self.symbol_var, width=20)
        self.symbol_entry.pack(fill=tk.X)
        
        # Quick filters
        ttk.Label(filter_frame, text="Quick Filters:", style='Header.TLabel').pack(anchor=tk.W, pady=(15, 5))
        
        ttk.Button(filter_frame, text="Top 10 Gainers", command=self.show_top_gainers).pack(fill=tk.X, pady=2)
        ttk.Button(filter_frame, text="Top 10 Losers", command=self.show_top_losers).pack(fill=tk.X, pady=2)
        ttk.Button(filter_frame, text="Biggest Trades", command=self.show_biggest_trades).pack(fill=tk.X, pady=2)
        ttk.Button(filter_frame, text="Most Active", command=self.show_most_active).pack(fill=tk.X, pady=2)
        
        # Apply/Reset buttons
        ttk.Separator(filter_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)
        
        ttk.Button(filter_frame, text="Apply Filters", command=self.apply_filters, 
                   style='Action.TButton').pack(fill=tk.X, pady=2)
        ttk.Button(filter_frame, text="Reset Filters", command=self.reset_filters).pack(fill=tk.X, pady=2)
        
    def create_raw_data_tab(self):
        """Create raw transactions data tab"""
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text="Raw Transactions")
        
        # Toolbar
        toolbar = ttk.Frame(tab)
        toolbar.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Button(toolbar, text="Refresh", command=self.refresh_raw_data).pack(side=tk.LEFT, padx=2)
        ttk.Button(toolbar, text="Export Selection", command=self.export_selection).pack(side=tk.LEFT, padx=2)
        
        # Record count label
        self.raw_count_label = ttk.Label(toolbar, text="Records: 0")
        self.raw_count_label.pack(side=tk.RIGHT)
        
        # Treeview with scrollbars
        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ['Date', 'Action', 'Symbol', 'Description', 'Qty', 'Price', 
                   'Gross', 'Net', 'Currency', 'Type']
        
        self.raw_tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
        
        # Configure columns
        col_widths = {'Date': 100, 'Action': 50, 'Symbol': 80, 'Description': 200, 
                      'Qty': 60, 'Price': 80, 'Gross': 100, 'Net': 100, 'Currency': 60, 'Type': 80}
        
        for col in columns:
            self.raw_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(self.raw_tree, c))
            self.raw_tree.column(col, width=col_widths.get(col, 100), anchor=tk.E if col in ['Qty', 'Price', 'Gross', 'Net'] else tk.W)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.raw_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.raw_tree.xview)
        self.raw_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.raw_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
    def create_trades_tab(self):
        """Create trades analysis tab"""
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text="Trades Analysis")
        
        # Summary frame at top
        summary_frame = ttk.LabelFrame(tab, text="Trade Summary", padding=10)
        summary_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.trade_summary_labels = {}
        metrics = ['Total Buys', 'Total Sells', 'Buy Value', 'Sell Value', 'Realized P&L']
        
        for i, metric in enumerate(metrics):
            ttk.Label(summary_frame, text=f"{metric}:").grid(row=0, column=i*2, padx=5, sticky=tk.E)
            label = ttk.Label(summary_frame, text="--", font=('Consolas', 10, 'bold'))
            label.grid(row=0, column=i*2+1, padx=(0, 20), sticky=tk.W)
            self.trade_summary_labels[metric] = label
        
        # Treeview
        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ['Symbol', 'Date', 'Action', 'Qty', 'Price', 'Value', 'Cost Basis', 'P&L', 'Category']
        
        self.trades_tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
        
        for col in columns:
            self.trades_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(self.trades_tree, c))
            self.trades_tree.column(col, width=100, anchor=tk.E if col in ['Qty', 'Price', 'Value', 'Cost Basis', 'P&L'] else tk.W)
        
        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.trades_tree.yview)
        self.trades_tree.configure(yscrollcommand=vsb.set)
        
        self.trades_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
    def create_dividends_tab(self):
        """Create dividends tab"""
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text="Dividends")
        
        # Summary
        summary_frame = ttk.LabelFrame(tab, text="Dividend Summary", padding=10)
        summary_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.div_summary_labels = {}
        for i, metric in enumerate(['Total Dividends', 'CAD Dividends', 'USD Dividends', 'Unique Stocks']):
            ttk.Label(summary_frame, text=f"{metric}:").grid(row=0, column=i*2, padx=5, sticky=tk.E)
            label = ttk.Label(summary_frame, text="--", font=('Consolas', 10, 'bold'))
            label.grid(row=0, column=i*2+1, padx=(0, 20), sticky=tk.W)
            self.div_summary_labels[metric] = label
        
        # Treeview
        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ['Date', 'Symbol', 'Description', 'Amount', 'Currency']
        
        self.div_tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
        
        for col in columns:
            self.div_tree.heading(col, text=col)
            self.div_tree.column(col, width=150 if col == 'Description' else 100)
        
        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.div_tree.yview)
        self.div_tree.configure(yscrollcommand=vsb.set)
        
        self.div_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
    def create_pnl_tab(self):
        """Create P&L summary tab"""
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text="P&L Summary")
        
        # Overall summary
        overall_frame = ttk.LabelFrame(tab, text="Overall Performance", padding=15)
        overall_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.pnl_summary_labels = {}
        metrics = ['Total Revenue', 'Total Cost', 'Realized P&L', 'Dividend Income', 'Total Return']
        
        for i, metric in enumerate(metrics):
            ttk.Label(overall_frame, text=f"{metric}:", font=('Segoe UI', 10)).grid(row=0, column=i*2, padx=10, sticky=tk.E)
            label = ttk.Label(overall_frame, text="--", font=('Consolas', 12, 'bold'))
            label.grid(row=0, column=i*2+1, padx=(0, 20), sticky=tk.W)
            self.pnl_summary_labels[metric] = label
        
        # Per-stock summary
        tree_frame = ttk.LabelFrame(tab, text="Per-Stock Summary", padding=5)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ['Symbol', 'Category', 'Trades', 'Revenue', 'Cost Basis', 'P&L', 'Dividends', 'Total Return', 'ROI %']
        
        self.pnl_tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
        
        for col in columns:
            self.pnl_tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(self.pnl_tree, c))
            width = 100 if col not in ['Category'] else 80
            self.pnl_tree.column(col, width=width, anchor=tk.E if col not in ['Symbol', 'Category'] else tk.W)
        
        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.pnl_tree.yview)
        self.pnl_tree.configure(yscrollcommand=vsb.set)
        
        self.pnl_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
    def create_charts_tab(self):
        """Create charts tab using simple canvas charts"""
        tab = ttk.Frame(self.notebook, padding=5)
        self.notebook.add(tab, text="Charts")
        
        # Chart selection
        chart_toolbar = ttk.Frame(tab)
        chart_toolbar.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(chart_toolbar, text="Chart Type:").pack(side=tk.LEFT, padx=5)
        self.chart_type_var = tk.StringVar(value="P&L by Category")
        chart_types = ["P&L by Category", "Top 10 Performers", "Top 10 Losers",
                       "Dividend Distribution", "Trades by Category"]
        self.chart_combo = ttk.Combobox(chart_toolbar, textvariable=self.chart_type_var,
                                        values=chart_types, state="readonly", width=25)
        self.chart_combo.pack(side=tk.LEFT, padx=5)
        self.chart_combo.bind('<<ComboboxSelected>>', lambda e: self.update_chart())
        
        ttk.Button(chart_toolbar, text="Refresh Chart", command=self.update_chart).pack(side=tk.LEFT, padx=10)
        
        # Chart canvas
        self.chart_frame = ttk.Frame(tab)
        self.chart_frame.pack(fill=tk.BOTH, expand=True)
        
        self.chart_canvas = SimpleBarChart(self.chart_frame, width=900, height=500)
        self.chart_canvas.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
    def create_status_bar(self):
        """Create status bar"""
        self.status_bar = ttk.Label(self.root, text="Ready - Import a file to begin", 
                                    relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
    def set_status(self, message):
        """Update status bar"""
        self.status_bar.config(text=message)
        self.root.update_idletasks()
        
    def import_excel(self):
        """Import Excel file"""
        filepath = filedialog.askopenfilename(
            title="Select Questrade Activities File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filepath:
            self.load_data(filepath, 'excel')
            
    def import_csv(self):
        """Import CSV file"""
        filepath = filedialog.askopenfilename(
            title="Select Questrade Activities File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filepath:
            self.load_data(filepath, 'csv')
            
    def load_data(self, filepath, file_type):
        """Load transaction data from file"""
        try:
            self.set_status(f"Loading {filepath}...")
            
            if file_type == 'excel':
                self.transactions_df = pd.read_excel(filepath)
            else:
                self.transactions_df = pd.read_csv(filepath)
            
            # Standardize column names
            self.transactions_df.columns = [col.strip().replace(' ', '_') for col in self.transactions_df.columns]
            
            # Parse dates
            if 'Transaction_Date' in self.transactions_df.columns:
                self.transactions_df['Transaction_Date'] = pd.to_datetime(
                    self.transactions_df['Transaction_Date'], format='mixed', errors='coerce'
                )
            
            # Add category column
            self.transactions_df['Category'] = self.transactions_df['Symbol'].apply(self.get_category)
            
            # Split trades and dividends
            self.trades_df = self.transactions_df[
                self.transactions_df['Activity_Type'] == 'Trades'
            ].copy()
            
            self.dividends_df = self.transactions_df[
                self.transactions_df['Activity_Type'] == 'Dividends'
            ].copy()
            
            # Calculate FIFO P&L
            self.calculate_fifo_pnl()
            
            # Update all views
            self.refresh_all_views()
            
            self.set_status(f"Loaded {len(self.transactions_df)} transactions from {os.path.basename(filepath)}")
            messagebox.showinfo("Success", f"Loaded {len(self.transactions_df)} transactions successfully!")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file:\n{str(e)}")
            self.set_status("Error loading file")
            
    def get_category(self, symbol):
        """Determine stock category"""
        if pd.isna(symbol):
            return "Other"
        symbol = str(symbol).upper()
        for category, symbols in self.categories.items():
            if symbol in [s.upper() for s in symbols]:
                return category
        return "Other"
        
    def calculate_fifo_pnl(self):
        """Calculate FIFO-based P&L for sells"""
        if self.trades_df is None or len(self.trades_df) == 0:
            return
            
        results = []
        symbols = self.trades_df['Symbol'].unique()
        
        for symbol in symbols:
            symbol_trades = self.trades_df[self.trades_df['Symbol'] == symbol].sort_values('Transaction_Date')
            
            # FIFO queue: list of (qty, price)
            lots = []
            
            for _, row in symbol_trades.iterrows():
                action = str(row.get('Action', '')).upper()
                qty = abs(float(row.get('Quantity', 0)))
                price = float(row.get('Price', 0))
                
                if action == 'BUY' and qty > 0:
                    lots.append({'qty': qty, 'price': price})
                    
                elif action == 'SELL' and qty > 0:
                    remaining = qty
                    cost_basis = 0
                    sell_revenue = qty * price
                    
                    while remaining > 0 and lots:
                        lot = lots[0]
                        take = min(remaining, lot['qty'])
                        cost_basis += take * lot['price']
                        lot['qty'] -= take
                        remaining -= take
                        
                        if lot['qty'] <= 0:
                            lots.pop(0)
                    
                    profit = sell_revenue - cost_basis
                    
                    results.append({
                        'Symbol': symbol,
                        'Date': row['Transaction_Date'],
                        'Quantity': qty,
                        'Sell_Price': price,
                        'Revenue': sell_revenue,
                        'Cost_Basis': cost_basis,
                        'Profit': profit,
                        'Category': self.get_category(symbol)
                    })
        
        self.fifo_results = pd.DataFrame(results) if results else pd.DataFrame()
        
        # Calculate per-stock summary
        self.calculate_stock_summary()
        
    def calculate_stock_summary(self):
        """Calculate per-stock summary"""
        if self.fifo_results is None or len(self.fifo_results) == 0:
            self.stock_summary = pd.DataFrame()
            return
            
        # Aggregate sells
        sell_summary = self.fifo_results.groupby('Symbol').agg({
            'Revenue': 'sum',
            'Cost_Basis': 'sum',
            'Profit': 'sum',
            'Quantity': 'count'
        }).rename(columns={'Quantity': 'Trades'})
        
        # Aggregate dividends
        if self.dividends_df is not None and len(self.dividends_df) > 0:
            div_summary = self.dividends_df.groupby('Symbol')['Net_Amount'].sum()
            # Clean up dividend symbols (remove leading dot)
            div_summary.index = div_summary.index.str.replace(r'^\.', '', regex=True)
            sell_summary['Dividends'] = sell_summary.index.map(lambda x: div_summary.get(x, 0))
        else:
            sell_summary['Dividends'] = 0
            
        sell_summary['Total_Return'] = sell_summary['Profit'] + sell_summary['Dividends']
        sell_summary['ROI'] = (sell_summary['Profit'] / sell_summary['Cost_Basis'] * 100).replace([np.inf, -np.inf], 0)
        sell_summary['Category'] = sell_summary.index.map(self.get_category)
        
        self.stock_summary = sell_summary.reset_index()
        
    def refresh_all_views(self):
        """Refresh all data views"""
        self.refresh_raw_data()
        self.refresh_trades_view()
        self.refresh_dividends_view()
        self.refresh_pnl_view()
        self.update_chart()
        
    def apply_filters(self):
        """Apply filters and refresh views"""
        self.refresh_all_views()
        
    def reset_filters(self):
        """Reset all filters"""
        self.date_from.delete(0, tk.END)
        self.date_from.insert(0, "2025-01-01")
        self.date_to.delete(0, tk.END)
        self.date_to.insert(0, "2025-12-31")
        self.category_var.set("All")
        self.action_var.set("All")
        self.currency_var.set("All")
        self.symbol_var.set("")
        self.refresh_all_views()
        
    def get_filtered_data(self, df):
        """Apply current filters to dataframe"""
        if df is None or len(df) == 0:
            return pd.DataFrame()
            
        filtered = df.copy()
        
        # Date filter
        try:
            date_from = pd.to_datetime(self.date_from.get())
            date_to = pd.to_datetime(self.date_to.get())
            if 'Transaction_Date' in filtered.columns:
                filtered = filtered[
                    (filtered['Transaction_Date'] >= date_from) & 
                    (filtered['Transaction_Date'] <= date_to)
                ]
        except:
            pass
            
        # Category filter
        if self.category_var.get() != "All":
            filtered = filtered[filtered['Category'] == self.category_var.get()]
            
        # Action filter
        if self.action_var.get() != "All":
            if 'Action' in filtered.columns:
                filtered = filtered[filtered['Action'] == self.action_var.get()]
                
        # Currency filter
        if self.currency_var.get() != "All":
            if 'Currency' in filtered.columns:
                filtered = filtered[filtered['Currency'] == self.currency_var.get()]
                
        # Symbol filter
        if self.symbol_var.get():
            if 'Symbol' in filtered.columns:
                filtered = filtered[filtered['Symbol'].str.contains(self.symbol_var.get(), case=False, na=False)]
                
        return filtered
        
    def refresh_raw_data(self):
        """Refresh raw data view"""
        self.raw_tree.delete(*self.raw_tree.get_children())
        
        if self.transactions_df is None:
            return
            
        filtered = self.get_filtered_data(self.transactions_df)
        
        for _, row in filtered.iterrows():
            values = (
                str(row.get('Transaction_Date', ''))[:10],
                row.get('Action', ''),
                row.get('Symbol', ''),
                str(row.get('Description', ''))[:40],
                row.get('Quantity', ''),
                f"${row.get('Price', 0):.2f}" if pd.notna(row.get('Price')) else '',
                f"${row.get('Gross_Amount', 0):,.2f}" if pd.notna(row.get('Gross_Amount')) else '',
                f"${row.get('Net_Amount', 0):,.2f}" if pd.notna(row.get('Net_Amount')) else '',
                row.get('Currency', ''),
                row.get('Activity_Type', '')
            )
            self.raw_tree.insert('', tk.END, values=values)
            
        self.raw_count_label.config(text=f"Records: {len(filtered):,}")
        
    def refresh_trades_view(self):
        """Refresh trades analysis view"""
        self.trades_tree.delete(*self.trades_tree.get_children())
        
        if self.fifo_results is None or len(self.fifo_results) == 0:
            return
            
        # Update summary
        total_buys = len(self.trades_df[self.trades_df['Action'].str.upper() == 'BUY']) if self.trades_df is not None else 0
        total_sells = len(self.fifo_results)
        buy_value = abs(self.trades_df[self.trades_df['Action'].str.upper() == 'BUY']['Net_Amount'].sum()) if self.trades_df is not None else 0
        sell_revenue = self.fifo_results['Revenue'].sum()
        realized_pnl = self.fifo_results['Profit'].sum()
        
        self.trade_summary_labels['Total Buys'].config(text=f"{total_buys:,}")
        self.trade_summary_labels['Total Sells'].config(text=f"{total_sells:,}")
        self.trade_summary_labels['Buy Value'].config(text=f"${buy_value:,.2f}")
        self.trade_summary_labels['Sell Value'].config(text=f"${sell_revenue:,.2f}")
        
        pnl_color = 'green' if realized_pnl >= 0 else 'red'
        self.trade_summary_labels['Realized P&L'].config(text=f"${realized_pnl:,.2f}", foreground=pnl_color)
        
        # Populate treeview
        for _, row in self.fifo_results.iterrows():
            values = (
                row['Symbol'],
                str(row['Date'])[:10],
                'Sell',
                f"{row['Quantity']:.0f}",
                f"${row['Sell_Price']:.2f}",
                f"${row['Revenue']:,.2f}",
                f"${row['Cost_Basis']:,.2f}",
                f"${row['Profit']:,.2f}",
                row['Category']
            )
            self.trades_tree.insert('', tk.END, values=values)
            
    def refresh_dividends_view(self):
        """Refresh dividends view"""
        self.div_tree.delete(*self.div_tree.get_children())
        
        if self.dividends_df is None or len(self.dividends_df) == 0:
            return
            
        filtered = self.get_filtered_data(self.dividends_df)
        
        # Update summary
        total_div = filtered['Net_Amount'].sum()
        cad_div = filtered[filtered['Currency'] == 'CAD']['Net_Amount'].sum()
        usd_div = filtered[filtered['Currency'] == 'USD']['Net_Amount'].sum()
        unique_stocks = filtered['Symbol'].nunique()
        
        self.div_summary_labels['Total Dividends'].config(text=f"${total_div:,.2f}")
        self.div_summary_labels['CAD Dividends'].config(text=f"${cad_div:,.2f}")
        self.div_summary_labels['USD Dividends'].config(text=f"${usd_div:,.2f}")
        self.div_summary_labels['Unique Stocks'].config(text=f"{unique_stocks}")
        
        for _, row in filtered.iterrows():
            values = (
                str(row.get('Transaction_Date', ''))[:10],
                row.get('Symbol', ''),
                str(row.get('Description', ''))[:50],
                f"${row.get('Net_Amount', 0):,.2f}",
                row.get('Currency', '')
            )
            self.div_tree.insert('', tk.END, values=values)
            
    def refresh_pnl_view(self):
        """Refresh P&L summary view"""
        self.pnl_tree.delete(*self.pnl_tree.get_children())
        
        if self.stock_summary is None or len(self.stock_summary) == 0:
            return
            
        # Overall summary
        total_revenue = self.stock_summary['Revenue'].sum()
        total_cost = self.stock_summary['Cost_Basis'].sum()
        total_pnl = self.stock_summary['Profit'].sum()
        total_div = self.stock_summary['Dividends'].sum()
        total_return = self.stock_summary['Total_Return'].sum()
        
        self.pnl_summary_labels['Total Revenue'].config(text=f"${total_revenue:,.2f}")
        self.pnl_summary_labels['Total Cost'].config(text=f"${total_cost:,.2f}")
        
        pnl_color = 'green' if total_pnl >= 0 else 'red'
        self.pnl_summary_labels['Realized P&L'].config(text=f"${total_pnl:,.2f}", foreground=pnl_color)
        self.pnl_summary_labels['Dividend Income'].config(text=f"${total_div:,.2f}")
        
        return_color = 'green' if total_return >= 0 else 'red'
        self.pnl_summary_labels['Total Return'].config(text=f"${total_return:,.2f}", foreground=return_color)
        
        # Per-stock table
        for _, row in self.stock_summary.iterrows():
            values = (
                row['Symbol'],
                row['Category'],
                f"{row['Trades']:.0f}",
                f"${row['Revenue']:,.2f}",
                f"${row['Cost_Basis']:,.2f}",
                f"${row['Profit']:,.2f}",
                f"${row['Dividends']:,.2f}",
                f"${row['Total_Return']:,.2f}",
                f"{row['ROI']:.1f}%"
            )
            self.pnl_tree.insert('', tk.END, values=values)
            
    def update_chart(self):
        """Update chart based on selection"""
        if self.stock_summary is None or len(self.stock_summary) == 0:
            self.chart_canvas.clear()
            self.chart_canvas.create_text(450, 250, text='No data available\nImport a file to view charts',
                                         font=('Segoe UI', 14), justify='center')
            return
            
        chart_type = self.chart_type_var.get()
        
        if chart_type == "P&L by Category":
            category_pnl = self.stock_summary.groupby('Category')['Profit'].sum().sort_values()
            self.chart_canvas.draw_bar_chart(
                category_pnl.values.tolist(),
                category_pnl.index.tolist(),
                title="Realized P&L by Category",
                horizontal=True
            )
            
        elif chart_type == "Top 10 Performers":
            top10 = self.stock_summary.nlargest(10, 'Profit')
            self.chart_canvas.draw_bar_chart(
                top10['Profit'].tolist(),
                top10['Symbol'].tolist(),
                title="Top 10 Performers by Profit"
            )
            
        elif chart_type == "Top 10 Losers":
            bottom10 = self.stock_summary.nsmallest(10, 'Profit')
            self.chart_canvas.draw_bar_chart(
                bottom10['Profit'].tolist(),
                bottom10['Symbol'].tolist(),
                title="Top 10 Losers by Profit"
            )
            
        elif chart_type == "Dividend Distribution":
            div_by_cat = self.stock_summary.groupby('Category')['Dividends'].sum()
            div_by_cat = div_by_cat[div_by_cat > 0]
            if len(div_by_cat) > 0:
                self.chart_canvas.draw_pie_chart(
                    div_by_cat.values.tolist(),
                    div_by_cat.index.tolist(),
                    title="Dividend Distribution by Category"
                )
            else:
                self.chart_canvas.clear()
                self.chart_canvas.create_text(450, 250, text='No dividend data',
                                             font=('Segoe UI', 14))
                
        elif chart_type == "Trades by Category":
            trades_by_cat = self.stock_summary.groupby('Category')['Trades'].sum().sort_values(ascending=False)
            self.chart_canvas.draw_bar_chart(
                trades_by_cat.values.tolist(),
                trades_by_cat.index.tolist(),
                title="Number of Trades by Category",
                colors='#2196F3'
            )
        
    def show_top_gainers(self):
        """Show top 10 gainers report"""
        if self.stock_summary is None or len(self.stock_summary) == 0:
            messagebox.showinfo("Info", "No data available. Please import a file first.")
            return
            
        top10 = self.stock_summary.nlargest(10, 'Profit')
        self.show_report_window("Top 10 Gainers", top10)
        
    def show_top_losers(self):
        """Show top 10 losers report"""
        if self.stock_summary is None or len(self.stock_summary) == 0:
            messagebox.showinfo("Info", "No data available. Please import a file first.")
            return
            
        bottom10 = self.stock_summary.nsmallest(10, 'Profit')
        self.show_report_window("Top 10 Losers", bottom10)
        
    def show_biggest_trades(self):
        """Show biggest trades by value"""
        if self.fifo_results is None or len(self.fifo_results) == 0:
            messagebox.showinfo("Info", "No trade data available.")
            return
            
        biggest = self.fifo_results.nlargest(10, 'Revenue')
        self.show_report_window("Biggest Trades by Value", biggest)
        
    def show_most_active(self):
        """Show most actively traded stocks"""
        if self.stock_summary is None or len(self.stock_summary) == 0:
            messagebox.showinfo("Info", "No data available.")
            return
            
        most_active = self.stock_summary.nlargest(10, 'Trades')
        self.show_report_window("Most Actively Traded Stocks", most_active)
        
    def show_by_category(self):
        """Show summary by category"""
        if self.stock_summary is None or len(self.stock_summary) == 0:
            messagebox.showinfo("Info", "No data available.")
            return
            
        category_summary = self.stock_summary.groupby('Category').agg({
            'Trades': 'sum',
            'Revenue': 'sum',
            'Cost_Basis': 'sum',
            'Profit': 'sum',
            'Dividends': 'sum',
            'Total_Return': 'sum'
        }).reset_index()
        
        self.show_report_window("Summary by Category", category_summary)
        
    def show_monthly_summary(self):
        """Show monthly summary"""
        if self.fifo_results is None or len(self.fifo_results) == 0:
            messagebox.showinfo("Info", "No data available.")
            return
            
        monthly = self.fifo_results.copy()
        monthly['Month'] = monthly['Date'].dt.to_period('M')
        monthly_summary = monthly.groupby('Month').agg({
            'Revenue': 'sum',
            'Cost_Basis': 'sum',
            'Profit': 'sum',
            'Quantity': 'count'
        }).rename(columns={'Quantity': 'Trades'}).reset_index()
        monthly_summary['Month'] = monthly_summary['Month'].astype(str)
        
        self.show_report_window("Monthly Summary", monthly_summary)
        
    def show_report_window(self, title, df):
        """Show report in a new window"""
        window = tk.Toplevel(self.root)
        window.title(title)
        window.geometry("900x500")
        
        # Treeview
        tree_frame = ttk.Frame(window, padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = list(df.columns)
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings')
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        
        for _, row in df.iterrows():
            values = []
            for col in columns:
                val = row[col]
                if isinstance(val, float):
                    values.append(f"${val:,.2f}" if abs(val) > 1 else f"{val:.4f}")
                else:
                    values.append(str(val))
            tree.insert('', tk.END, values=values)
        
        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Export button
        btn_frame = ttk.Frame(window, padding=10)
        btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="Export to Excel", 
                   command=lambda: self.quick_export_excel(df, title)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Close", command=window.destroy).pack(side=tk.RIGHT, padx=5)
        
    def quick_export_excel(self, df, title):
        """Quick export dataframe to Excel"""
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"{title.replace(' ', '_')}.xlsx"
        )
        if filepath:
            df.to_excel(filepath, index=False)
            messagebox.showinfo("Success", f"Exported to {filepath}")
        
    def sort_treeview(self, tree, col):
        """Sort treeview by column"""
        items = [(tree.set(k, col), k) for k in tree.get_children('')]
        try:
            items.sort(key=lambda t: float(t[0].replace('$', '').replace(',', '').replace('%', '')))
        except:
            items.sort()
        for index, (val, k) in enumerate(items):
            tree.move(k, '', index)
            
    def show_tab(self, index):
        """Switch to specific tab"""
        self.notebook.select(index)
        
    def export_excel(self):
        """Export full report to Excel"""
        if self.transactions_df is None:
            messagebox.showinfo("Info", "No data to export. Please import a file first.")
            return
            
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Trading_Report.xlsx"
        )
        
        if not filepath:
            return
            
        try:
            self.set_status("Exporting to Excel...")
            
            wb = Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4472C4")
            
            # Write summary
            ws_summary['A1'] = "Trading Report Summary"
            ws_summary['A1'].font = Font(bold=True, size=16)
            ws_summary.merge_cells('A1:D1')
            
            ws_summary['A3'] = "Generated:"
            ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if self.stock_summary is not None:
                ws_summary['A4'] = "Total Realized P&L:"
                ws_summary['B4'] = self.stock_summary['Profit'].sum()
                ws_summary['A5'] = "Total Dividends:"
                ws_summary['B5'] = self.stock_summary['Dividends'].sum()
                ws_summary['A6'] = "Total Return:"
                ws_summary['B6'] = self.stock_summary['Total_Return'].sum()
            
            # Transactions sheet
            ws_trans = wb.create_sheet("Transactions")
            for r_idx, row in enumerate(dataframe_to_rows(self.transactions_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_trans.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
            
            # P&L sheet
            if self.stock_summary is not None and len(self.stock_summary) > 0:
                ws_pnl = wb.create_sheet("P&L Summary")
                for r_idx, row in enumerate(dataframe_to_rows(self.stock_summary, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws_pnl.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
            
            # Dividends sheet
            if self.dividends_df is not None and len(self.dividends_df) > 0:
                ws_div = wb.create_sheet("Dividends")
                for r_idx, row in enumerate(dataframe_to_rows(self.dividends_df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws_div.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
            
            # FIFO Sells sheet
            if self.fifo_results is not None and len(self.fifo_results) > 0:
                ws_fifo = wb.create_sheet("FIFO Sells")
                for r_idx, row in enumerate(dataframe_to_rows(self.fifo_results, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws_fifo.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
            
            wb.save(filepath)
            self.set_status(f"Exported to {filepath}")
            messagebox.showinfo("Success", f"Report exported to:\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export:\n{str(e)}")
            
    def export_pdf(self):
        """Export report to PDF"""
        if self.transactions_df is None:
            messagebox.showinfo("Info", "No data to export.")
            return
            
        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile="Trading_Report.pdf"
        )
        
        if not filepath:
            return
            
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            
            self.set_status("Exporting to PDF...")
            
            doc = SimpleDocTemplate(filepath, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1)
            elements.append(Paragraph("Trading Report", title_style))
            elements.append(Spacer(1, 20))
            
            # Summary
            elements.append(Paragraph("Summary", styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            if self.stock_summary is not None:
                summary_data = [
                    ["Metric", "Value"],
                    ["Total Realized P&L", f"${self.stock_summary['Profit'].sum():,.2f}"],
                    ["Total Dividends", f"${self.stock_summary['Dividends'].sum():,.2f}"],
                    ["Total Return", f"${self.stock_summary['Total_Return'].sum():,.2f}"],
                    ["Number of Stocks", str(len(self.stock_summary))],
                ]
                
                summary_table = Table(summary_data)
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(summary_table)
                elements.append(Spacer(1, 20))
            
            # P&L by Stock (top 20)
            if self.stock_summary is not None and len(self.stock_summary) > 0:
                elements.append(Paragraph("P&L by Stock (Top 20)", styles['Heading2']))
                elements.append(Spacer(1, 10))
                
                top20 = self.stock_summary.nlargest(20, 'Total_Return')
                pnl_data = [['Symbol', 'Category', 'Trades', 'Revenue', 'Cost', 'P&L', 'Dividends', 'Total']]
                
                for _, row in top20.iterrows():
                    pnl_data.append([
                        str(row['Symbol']),
                        str(row['Category']),
                        str(int(row['Trades'])),
                        f"${row['Revenue']:,.0f}",
                        f"${row['Cost_Basis']:,.0f}",
                        f"${row['Profit']:,.0f}",
                        f"${row['Dividends']:,.0f}",
                        f"${row['Total_Return']:,.0f}"
                    ])
                
                pnl_table = Table(pnl_data)
                pnl_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                elements.append(pnl_table)
            
            doc.build(elements)
            self.set_status(f"Exported to {filepath}")
            messagebox.showinfo("Success", f"PDF exported to:\n{filepath}")
            
        except ImportError:
            messagebox.showerror("Error", "ReportLab library required for PDF export.\nInstall with: pip install reportlab")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export PDF:\n{str(e)}")
            
    def export_html(self):
        """Export report to HTML"""
        if self.transactions_df is None:
            messagebox.showinfo("Info", "No data to export.")
            return
            
        filepath = filedialog.asksaveasfilename(
            defaultextension=".html",
            filetypes=[("HTML files", "*.html")],
            initialfile="Trading_Report.html"
        )
        
        if not filepath:
            return
            
        try:
            self.set_status("Exporting to HTML...")
            
            # Generate HTML
            html_content = self.generate_html_report()
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            self.set_status(f"Exported to {filepath}")
            
            # Ask to open in browser
            if messagebox.askyesno("Success", f"HTML exported to:\n{filepath}\n\nOpen in browser?"):
                webbrowser.open(f'file://{os.path.abspath(filepath)}')
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export HTML:\n{str(e)}")
            
    def generate_html_report(self):
        """Generate HTML report content"""
        total_pnl = self.stock_summary['Profit'].sum() if self.stock_summary is not None else 0
        total_div = self.stock_summary['Dividends'].sum() if self.stock_summary is not None else 0
        total_return = self.stock_summary['Total_Return'].sum() if self.stock_summary is not None else 0
        
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Trading Report</title>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; background: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 15px; }}
        h2 {{ color: #34495e; margin-top: 30px; }}
        .summary-cards {{ display: flex; gap: 20px; flex-wrap: wrap; margin: 20px 0; }}
        .card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; min-width: 200px; flex: 1; }}
        .card.positive {{ background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); }}
        .card.negative {{ background: linear-gradient(135deg, #eb3349 0%, #f45c43 100%); }}
        .card h3 {{ margin: 0 0 10px 0; font-size: 14px; opacity: 0.9; }}
        .card .value {{ font-size: 24px; font-weight: bold; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th {{ background: #3498db; color: white; padding: 12px; text-align: left; }}
        td {{ padding: 10px; border-bottom: 1px solid #eee; }}
        tr:hover {{ background: #f8f9fa; }}
        .positive {{ color: #27ae60; }}
        .negative {{ color: #e74c3c; }}
        .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #eee; color: #7f8c8d; font-size: 12px; }}
        @media print {{ body {{ margin: 0; }} .container {{ box-shadow: none; }} }}
    </style>
</head>
<body>
    <div class="container">
        <h1>📈 Trading Report</h1>
        <p>Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
        
        <div class="summary-cards">
            <div class="card {'positive' if total_pnl >= 0 else 'negative'}">
                <h3>Realized P&L</h3>
                <div class="value">${total_pnl:,.2f}</div>
            </div>
            <div class="card">
                <h3>Dividend Income</h3>
                <div class="value">${total_div:,.2f}</div>
            </div>
            <div class="card {'positive' if total_return >= 0 else 'negative'}">
                <h3>Total Return</h3>
                <div class="value">${total_return:,.2f}</div>
            </div>
            <div class="card">
                <h3>Stocks Traded</h3>
                <div class="value">{len(self.stock_summary) if self.stock_summary is not None else 0}</div>
            </div>
        </div>
        
        <h2>P&L by Stock</h2>
        <table>
            <thead>
                <tr>
                    <th>Symbol</th>
                    <th>Category</th>
                    <th>Trades</th>
                    <th>Revenue</th>
                    <th>Cost Basis</th>
                    <th>P&L</th>
                    <th>Dividends</th>
                    <th>Total Return</th>
                </tr>
            </thead>
            <tbody>
"""
        
        if self.stock_summary is not None:
            for _, row in self.stock_summary.sort_values('Total_Return', ascending=False).iterrows():
                pnl_class = 'positive' if row['Profit'] >= 0 else 'negative'
                total_class = 'positive' if row['Total_Return'] >= 0 else 'negative'
                html += f"""
                <tr>
                    <td><strong>{row['Symbol']}</strong></td>
                    <td>{row['Category']}</td>
                    <td>{int(row['Trades'])}</td>
                    <td>${row['Revenue']:,.2f}</td>
                    <td>${row['Cost_Basis']:,.2f}</td>
                    <td class="{pnl_class}">${row['Profit']:,.2f}</td>
                    <td>${row['Dividends']:,.2f}</td>
                    <td class="{total_class}"><strong>${row['Total_Return']:,.2f}</strong></td>
                </tr>
"""
        
        html += """
            </tbody>
        </table>
        
        <div class="footer">
            <p>Report generated by Trading Report Builder</p>
        </div>
    </div>
</body>
</html>
"""
        return html
        
    def print_report(self):
        """Print report (generates HTML and opens print dialog)"""
        if self.transactions_df is None:
            messagebox.showinfo("Info", "No data to print.")
            return
            
        try:
            # Generate temporary HTML file
            html_content = self.generate_html_report()
            
            with tempfile.NamedTemporaryFile(mode='w', suffix='.html', delete=False, encoding='utf-8') as f:
                f.write(html_content)
                temp_path = f.name
            
            # Open in browser for printing
            webbrowser.open(f'file://{temp_path}')
            messagebox.showinfo("Print", "Report opened in browser.\nUse Ctrl+P to print.")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate print preview:\n{str(e)}")
            
    def export_selection(self):
        """Export selected rows"""
        selected = self.raw_tree.selection()
        if not selected:
            messagebox.showinfo("Info", "No rows selected")
            return
            
        # Get selected data
        data = []
        for item in selected:
            data.append(self.raw_tree.item(item)['values'])
            
        df = pd.DataFrame(data, columns=['Date', 'Action', 'Symbol', 'Description', 'Qty', 
                                         'Price', 'Gross', 'Net', 'Currency', 'Type'])
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        )
        
        if filepath:
            if filepath.endswith('.csv'):
                df.to_csv(filepath, index=False)
            else:
                df.to_excel(filepath, index=False)
            messagebox.showinfo("Success", f"Exported {len(data)} rows to {filepath}")
            
    def show_about(self):
        """Show about dialog"""
        messagebox.showinfo("About", 
            "Trading Report Builder\n"
            "Version 1.1 (Portable)\n\n"
            "A desktop application for analyzing\n"
            "Questrade trading transactions.\n\n"
            "Features:\n"
            "• Import Excel/CSV transactions\n"
            "• FIFO P&L calculations\n"
            "• Charts and visualizations\n"
            "• Export to Excel, PDF, HTML\n"
            "• Print reports"
        )


def main():
    root = tk.Tk()
    app = TradingReportBuilder(root)
    root.mainloop()


if __name__ == "__main__":
    main()
